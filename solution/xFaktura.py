import collections
import datetime
import glob
import locale
import numpy as np
import os
import pandas as pd
import platform
import re
import shutil
import subprocess
import sys
import time
import watchdog.events
import watchdog.observers
from openpyxl import load_workbook

print(f"xFaktura 1.7.0, Python {platform.python_version()}, {platform.system()} {platform.release()}")


# EN headers of spreadsheet invoices
spreadsheet_invoices_headers = {}
spreadsheet_invoices_headers['invoice']           = 'Invoice'
spreadsheet_invoices_headers['date_invoice']      = 'Date invoice'
spreadsheet_invoices_headers['salutation']        = 'Salutation'
spreadsheet_invoices_headers['first_name']        = 'First name'
spreadsheet_invoices_headers['last_name']         = 'Last name'
spreadsheet_invoices_headers['street']            = 'Street'
spreadsheet_invoices_headers['city']              = 'City'
spreadsheet_invoices_headers['physician']         = 'Physician'
spreadsheet_invoices_headers['date_prescription'] = 'Date Prescription'
spreadsheet_invoices_headers['diagnosis']         = 'Diagnosis'


# EN user feedback
Please_leave_only_one_tex_file_here_found= 'Please leave only one .tex template, found'
Please_leave_only_one_xls_file_here_found= 'Please leave only one .xls file, found'
Please_store_one_xls_file_here = 'Please store one .xls file here'
Please_remove_the_duplicated_header = 'Please remove the duplicated header'
Please_add_the_missing_header = 'Please add the missing header'
Found = 'Found'
Watch_out_for_spaces = 'Watch out for spaces'
Skipping_invoice_1_because_it_has_no_date = 'Skipping invoice {} because it has no date'


def set_language(LANG):
    global spreadsheet_invoices_headers
    global Please_leave_only_one_tex_file_here_found
    global Please_leave_only_one_xls_file_here_found
    global Please_store_one_xls_file_here
    global Please_remove_the_duplicated_header
    global Please_add_the_missing_header
    global Found
    global Watch_out_for_spaces
    global Skipping_invoice_1_because_it_has_no_date
    if LANG == 'de':
        locale.setlocale(locale.LC_ALL, 'de_DE.UTF-8')  # Voraussetzung für Komma bei Zahlen
        # DE headers of spreadsheet invoices
        spreadsheet_invoices_headers['invoice']           = 'Rechnung'
        spreadsheet_invoices_headers['date_invoice']      = 'Datum Rechnung'
        spreadsheet_invoices_headers['salutation']        = 'Anrede'
        spreadsheet_invoices_headers['first_name']        = 'Vorname'
        spreadsheet_invoices_headers['last_name']         = 'Nachname'
        spreadsheet_invoices_headers['street']            = 'Straße'
        spreadsheet_invoices_headers['city']              = 'Stadt'
        spreadsheet_invoices_headers['physician']         = 'Arzt'
        spreadsheet_invoices_headers['date_prescription'] = 'Datum Verordnung'
        spreadsheet_invoices_headers['diagnosis']         = 'Diagnose'
        # DE user feedback
        Please_leave_only_one_tex_file_here_found= 'Bitte nur eine .tex Vorlage, gefunden'
        Please_leave_only_one_xls_file_here_found= 'Bitte nur eine .xlxs Datei, gefunden'
        Please_store_one_xls_file_here = 'Bitte speichere eine Excel Datei hier'
        Please_remove_the_duplicated_header = 'Bitte entferne eine der doppelten Header'
        Please_add_the_missing_header = 'Bitte ergänze den fehlenden Header'
        Found = 'Gefunden'
        Watch_out_for_spaces = 'Beachte Leerzeichen'
        Skipping_invoice_1_because_it_has_no_date = 'Rechnung {} übersprungen weil Datum fehlt'



operatingSystem = platform.system()

def findExecutables():
    def couldbe(fpath):
        if os.path.isfile(fpath):
            return fpath
        else:
            return None

    if operatingSystem == 'Linux' or operatingSystem == 'Darwin':
        lualatex = shutil.which('lualatex')
        dvipdfmx = shutil.which('dvipdfmx')
        # Livetex may not install to path, therefore search at known locations
        if lualatex is None:
            for file in glob.iglob('/usr/local/texlive/**/bin/**/lualatex', recursive=True):
                lualatex = file
        if dvipdfmx is None:
            for file in glob.iglob('/usr/local/texlive/**/bin/**/dvipdfmx', recursive=True):
                dvipdfmx = file
        if operatingSystem == 'Darwin':
            # Livetex does not install to path, therefore search at known location
            if lualatex is None:
                lualatex = couldbe('/Library/TeX/texbin/lualatex')
            if dvipdfmx is None:
                dvipdfmx = couldbe('/Library/TeX/texbin/dvipdfmx')
    else:
        print('ERROR    system not covered')
        sys.exit(1)
    return lualatex, dvipdfmx




lualatex, dvipdfmx = findExecutables()
if lualatex is None or dvipdfmx is None:
    print('ERROR lualatex missing, please install')
    sys.exit(2)


try:
    tex_process = subprocess.run([ lualatex, '--version' ], capture_output=True)
    tex_out = tex_process.stdout.decode("utf-8").splitlines()[0]
    regexMatch = re.search('Version ([\d\.]+ \(TeX Live [\d]+\))', tex_out)
    if regexMatch:
        tex_version = regexMatch.group(1)
except:
    tex_version = '?'
print( f'Lualatex {tex_version}')
print( f'Zum Erzeugen der Rechnungen Excel speichern.')
print( f'')




# Chose TeX Template
TeXtemplateFiles = glob.glob("*.tex")
if len(TeXtemplateFiles) == 2 and 'Praxis1-Vorlage.tex' in TeXtemplateFiles:
    TeXtemplateFiles.remove('Praxis1-Vorlage.tex')
if len(TeXtemplateFiles) == 1:
    TeXtemplateFile = TeXtemplateFiles[0]
    TeXtemplateBasename = TeXtemplateFile.replace('.tex', '')
else:
    print(f"{Please_leave_only_one_tex_file_here_found} {TeXtemplateFiles}")
    sys.exit(1)


# Collect all-caps from TeX
allcap = ''
allcapDict = {}
with open(TeXtemplateFile, encoding='utf8') as file:
    for line in file:
        for char in line:
            if char.isupper() or char in ['ß', '_']:
                allcap += char
            else:
                if len(allcap) > 3:
                    allcapDict[allcap] = 1
                allcap = ''


if 'RECHNUNGSNUMMER' in allcapDict:
    set_language('de')


# Chose xls
xlsFiles = glob.glob("*.xlsx")
if len(xlsFiles) == 2 and 'Praxis1.xlsx' in xlsFiles:
    xlsFiles.remove('Praxis1.xlsx')
if len(xlsFiles) == 1:
    xlsFile = xlsFiles[0]
elif len(xlsFiles) == 0:
    print(f"{Please_store_one_xls_file_here}")
    sys.exit(1)
else:
    print(f"{Please_leave_only_one_xls_file_here_found} {xlsFiles}")
    sys.exit(1)


# Inspect data
Exceldatei = load_workbook(xlsFile)
Exceltabelle_Behandlungen = Exceldatei['Behandlungen']
Exceltabelle_Rechnungen = Exceldatei['Rechnungen']

# Inspect data -- headers
headersB = []
for header in [Exceltabelle_Behandlungen.cell(row=1,column=ccc) for ccc in range(1,40)]:
    if header.value is not None:
        headersB.append(header.value)

duplicatedB = [item for item, count in collections.Counter(headersB).items() if count > 1]
if len(duplicatedB) > 0:
    print(f"{Please_remove_the_duplicated_header} {duplicatedB}")
    sys.exit(1)

for muss in ['Rechnung', 'Behandlung', 'Patient', 'Rechnung']:
    if muss not in headersB:
        print(f"{Please_add_the_missing_header} {muss}")
        print(f"{Found} {headersB}")
        print(Watch_out_for_spaces)
        sys.exit(1)

headersR = []
for header in [Exceltabelle_Rechnungen.cell(row=1,column=ccc) for ccc in range(1,40)]:
    if header.value is not None:
        headersR.append(header.value)

for _,spreadsheet_invoices_header in spreadsheet_invoices_headers.items():
    if spreadsheet_invoices_header not in headersR:
        print(f"{Please_add_the_missing_header} {spreadsheet_invoices_header}")
        print(f"{Found} {headersR}")
        print(Watch_out_for_spaces)
        sys.exit(1)


# DataFrame ist ein Array aus Sheet, Column, Row. Leere Zellen, auch Text, werden zu float NaN
df_sheets = pd.read_excel(xlsFile, sheet_name=None)
Anzahl_pdf_nicht_überschrieben = 0
Anzahl_pdf_geschrieben = 0

aSheet = df_sheets["Behandlungen"]
#print(f"markus {aSheet.columns.values}")


def format_datet(datum_oder_text):
    if   type(datum_oder_text) == type(None):                    return ''
    elif type(datum_oder_text) == type(str()):                   return datum_oder_text
    elif type(datum_oder_text) == type(pd.Timestamp.now()):      return datum_oder_text.strftime('%d.%m.%Y')
    elif type(datum_oder_text) == type(datetime.datetime.now()): return datum_oder_text.strftime('%d.%m.%Y')
    else: return str(datum_oder_text)


def escape_latex_special_characters(in11):
    out11 = in11
    out11 = out11.replace('\\', '\\\\textbackslash')
    out11 = out11.replace('~', '\\\\textasciitilde')
    out11 = out11.replace('^', '\\\\textasciicircum')
    out11 = out11.replace('{', '\\{')
    out11 = out11.replace('}', '\\}')
    out11 = out11.replace('&', '\\&')
    out11 = out11.replace('%', '\\%')
    out11 = out11.replace('$', '\\$')
    out11 = out11.replace('#', '\\#')
    out11 = out11.replace('_', '\\_')
    return out11


def format_textt(text_oder_float_nan):
    # pandas gibt float nan als Leere Zelle zurück
    if type(text_oder_float_nan) == type(None): return ''
    elif type(text_oder_float_nan) == type(str()): return escape_latex_special_characters(text_oder_float_nan.rstrip())
    elif type(text_oder_float_nan) == type(1.0):
        if np.isnan(text_oder_float_nan):
            return ''
        else:
            return str(text_oder_float_nan)
    else: return str(text_oder_float_nan)


def write_error(txt):
    print(txt)
    #quit()


def lösche_datei(datei):
    if os.path.exists(datei):
        os.remove(datei)



def HACKnan(aaa):
    # TODO learn np.isnan
    return aaa != aaa



try:
    ExcelLine = 2
    for rechnr in df_sheets["Behandlungen"].Rechnung:
        #print(f" Zeile {ExcelLine} {rechnr}")
        if HACKnan(rechnr):
            print('')
            print(f"Behandlungen Zeile {ExcelLine} ist leer")
            print(f"Bitte leere Spalten löschen")
            sys.exit(1)
        if not re.match('^20\d\d-\d\d\d$', rechnr):
            write_error(f'Die Rechnungsnummer "{rechnr}" auf dem Blatt Behandlungen Zeile {ExcelLine} hat nicht das Format 20##-###')
            print(f"Bitte korrigieren")
            sys.exit(1)
        ExcelLine += 1


    ExcelLine = 2
    for rechnr in df_sheets["Rechnungen"].Rechnung:
        #print(f" Zeile {ExcelLine} {rechnr}")
        if HACKnan(rechnr):
            print('')
            print(f"Rechnungen Zeile {ExcelLine} ist leer")
            print(f"Bitte leere Spalten löschen")
            sys.exit(1)
        if not re.match('^20\d\d-\d\d\d$', rechnr):
            write_error(f'Die Rechnungsnummer "{rechnr}" auf dem Blatt Rechnungen Zeile {ExcelLine} hat nicht das Format 20##-###')
            print(f"Bitte korrigieren")
            sys.exit(1)
        ExcelLine += 1



    Rechnungsnummern = sorted(df_sheets["Behandlungen"].Rechnung.unique())
    if len(Rechnungsnummern) == 0:
        write_error('Keine Rechnungsnummern')

except KeyError as exc:
    write_error(f'In Excel fehlt das Blatt oder die Spalte {exc}')


def Diese_Rechnung(Rechnungsnummer):
    global Anzahl_pdf_nicht_überschrieben
    global Anzahl_pdf_geschrieben

    pdfglob  = f'Pdf/{TeXtemplateBasename}-{Rechnungsnummer}-*.pdf'
    if len(glob.glob(pdfglob)) > 0:
        #print(f"{pdfglob} nicht überschrieben")
        Anzahl_pdf_nicht_überschrieben += 1
        return

    #print(f'...{Rechnungsnummer}...', end='')
    date_invoice          = "fehlt"
    salutation            = "fehlt"
    first_name            = "fehlt"
    last_name             = "fehlt"
    street                = "fehlt"
    city                  = "fehlt"
    physician             = None
    date_prescription     = None
    diagnosis             = None
    BEHANDLUNGEN          = ''
    RECHNUNGSBETRAG_zahl  = 0
    treatements_count = []
    treatements_name  = []
    treatements_price = []
    treatements_sum   = []

    try:
        # Sammele Rechnungsdaten
        # Select rows where invoice column value equals incoice number given in function
        rechnung_df = df_sheets["Rechnungen"][(df_sheets["Rechnungen"]['Rechnung'] == Rechnungsnummer)]
        if rechnung_df['Rechnung'].size == 0:
            write_error(f'Rechnungsnummer {Rechnungsnummer} fehlt im Blatt Rechnungen. Sind Leerzeichen am Ende?')
            return
        if rechnung_df['Rechnung'].size != 1:
            write_error(f'Rechnungsnummer {Rechnungsnummer} gibt es im Blatt Rechnungen {rechnung_df.size}-fach')
            return
        first_name        = format_textt(rechnung_df[spreadsheet_invoices_headers['first_name']].item())    # LEARN DataFrame Daten sind items
        last_name         = format_textt(rechnung_df[spreadsheet_invoices_headers['last_name']].item())
        street            = format_textt(rechnung_df[spreadsheet_invoices_headers['street']].item())
        city              = format_textt(rechnung_df[spreadsheet_invoices_headers['city']].item())
        salutation        = format_textt(rechnung_df[spreadsheet_invoices_headers['salutation']].item())
        diagnosis         = format_textt(rechnung_df[spreadsheet_invoices_headers['diagnosis']].item())
        physician         = format_textt(rechnung_df[spreadsheet_invoices_headers['physician']].item())
        date_invoice      = format_datet(rechnung_df[spreadsheet_invoices_headers['date_invoice']].item())
        date_prescription = format_datet(rechnung_df[spreadsheet_invoices_headers['date_prescription']].item())

        # Inspect data -- date naT TODO
        if len(date_invoice) < 5:
            print(Skipping_invoice_1_because_it_has_no_date.format(Rechnungsnummer))
            return
        # Suche Behandlungsarten
        Behandlungsarten = []
        headersAb2 = df_sheets["Behandlungen"].columns[2:]
        # Beginne hinter Rechnung oder Rechnung.1 (Name der Spalte mehrfach)
        # Starte ab der 2. Spalte
        # Header vor Rechnung und ab Ausgaben aulassen
        header_auslassen = True
        for header in headersAb2:
            if header.startswith('Ausgaben'): header_auslassen = True
            if not header_auslassen: Behandlungsarten.append(header)
            if header.startswith('Rechnung'): header_auslassen = False
        #print(f'{Behandlungsarten=}')

        # Dataframe dieser Rechnungsnummer, mit allen Behandlungsarten
        behandlungen_df = df_sheets["Behandlungen"][(df_sheets["Behandlungen"]['Rechnung'] == Rechnungsnummer)]

        BEHANDLUNGSTERMINE_list_of_string = []
        for datum_oder_text in behandlungen_df['Behandlung']:
            BEHANDLUNGSTERMINE_list_of_string.append(format_datet(datum_oder_text))
        BEHANDLUNGSTERMINE = ', '.join(BEHANDLUNGSTERMINE_list_of_string)

        # Pro Behandlungsart
        # ANZAHL & BEHANLDUNG &  EINZELPREIS\,€ &  GESAMTPREIS\,€ \\
        for Behandlungsart in Behandlungsarten:
            df_behandlung = behandlungen_df[Behandlungsart]
            if df_behandlung.count() > 0:
                Anzahl      = df_behandlung.count()
                Einzelpreis = locale.format_string('%.2f', df_behandlung.mean())
                Gesamtpreis = locale.format_string('%.2f', df_behandlung.sum())
                treatements_count.append(str(Anzahl))
                treatements_name.append(Behandlungsart)
                treatements_price.append(Einzelpreis)
                treatements_sum.append(Gesamtpreis)
                # Das geht durch ein replace,  dafür muß ich \ escapen
                BEHANDLUNGEN += fr'  {Anzahl} & {Behandlungsart} & {Einzelpreis}\\,€ & {Gesamtpreis}\\,€ \\\\' + '\n'
                RECHNUNGSBETRAG_zahl += df_behandlung.sum()
        RECHNUNGSBETRAG = locale.format_string('%.2f', RECHNUNGSBETRAG_zahl)
    except KeyError as exc:
        write_error(f'In Excel fehlt das Blatt oder die Spalte {exc}')
        return


    # Benutze Vorlage und ersetze Werte
    output = ''
    inside_document = False
    kopie = ''
    with open(TeXtemplateFile, encoding='utf8') as file:
        for line in file:
            # Füge (Kopie) in der Kopie hinzu
            if r'\textbf{RECHNUNG}' in line:
                für_ORIG  = re.sub(r'\\textbf{RECHNUNG}', r'\\textbf{Rechnung}', line)
                für_KOPIE = re.sub(r'\\textbf{RECHNUNG}', r'\\textbf{Rechnung} (Kopie)', line)
                output += für_ORIG
                kopie += für_KOPIE
                continue

            # Ersetze die Platzhalter
            #         ANZAHL & BEHANDLUNG & EINZELPREIS\,€ & GESAMTPREIS\,€ \\
            line = re.sub(r'\bRECHNUNGSNUMMER\b', Rechnungsnummer, line)
            line = re.sub(r'\bRECHNUNGSDATUM\b', date_invoice, line)
            line = re.sub(r'\bVERORDNUNGSDATUM\b', date_prescription, line)
            line = re.sub(r'\bDIAGNOSE\b', diagnosis, line)
            line = re.sub(r'\bARZT\b', physician, line)
            line = re.sub(r'\bANREDE\b', salutation, line)
            line = re.sub(r'\bVORNAME\b', first_name, line)
            line = re.sub(r'\bNACHNAME\b', last_name, line)
            line = re.sub(r'\bSTRAßE\b', street, line)
            line = re.sub(r'\bSTADT\b', city, line)
            line = re.sub(r'\bRECHNUNGSBETRAG\b', RECHNUNGSBETRAG, line)
            ###line = re.sub(r'ANZAHL & BEHANDLUNG & EINZELPREIS\\,€ & GESAMTPREIS\\,€ \\\\', BEHANDLUNGEN, line)
            if 'ANZAHL' in line and 'BEHANDLUNG' in line and 'GESAMTPREIS' in line:
                line0 = line
                line2 = ''
                for idx in range(0, len(treatements_count)):
                    line1 = line0
                    line1 = re.sub(r'\bANZAHL\b',      treatements_count[idx],         line1)
                    line1 = re.sub(r'\bBEHANDLUNG\b',  treatements_name[idx], line1)
                    line1 = re.sub(r'\bEINZELPREIS\b', treatements_price[idx],    line1)
                    line1 = re.sub(r'\bGESAMTPREIS\b', treatements_sum[idx],    line1)
                    line2 += line1
                line = line2
            line = re.sub(r'\bBEHANDLUNGSTERMINE\b', BEHANDLUNGSTERMINE, line)

            # Erstelle Kopie
            if r'\end{document}' in line:
                output += r'\newpage'
                output += kopie
            output += line
            if inside_document:
                kopie += line
            if r'\begin{document}' in line:
                inside_document = True


    import platform
    tmpdir = '/tmp/xFaktura/'
    if platform.system() == 'Windows':
        tmpdir = 'C:\\Windows\\temp\\xFaktura\\'

    # https://csatlas.com/python-create-directory
    import os
    import sys
    if sys.version_info[1] > 4 or (sys.version_info[1] == 4 and sys.version_info[1] >= 1):
        os.makedirs(tmpdir, exist_ok=True)
        os.makedirs('Pdf', exist_ok=True)    # Store for pdf
    else:
        try:
            os.makedirs(tmpdir)
            os.makedirs('Pdf')
        except OSError:
            if not os.path.isdir(tmpdir) or not os.path.isdir('Pdf'):
                raise

    Basisname_der_Datei = f'{TeXtemplateBasename}-{Rechnungsnummer}-{last_name}'
    pdfdatei = 'Pdf/' + Basisname_der_Datei + '.pdf'
    texdatei = tmpdir + Basisname_der_Datei + '.tex'
    dvidatei = tmpdir + Basisname_der_Datei + '.dvi'
    auxdatei = tmpdir + Basisname_der_Datei + '.aux'
    logdatei = tmpdir + Basisname_der_Datei + '.log'

    with open(texdatei, 'w', encoding='utf8') as fout:
        fout.write(output)

    #print(f'{lualatex} --interaction=nonstopmode -output-directory={tmpdir} -output-format=dvi {texdatei}')

    tex_process = subprocess.run([ lualatex, '--interaction=nonstopmode', '-output-directory='+tmpdir, '-output-format=dvi', texdatei ], capture_output=True)
    if tex_process.returncode == 0:
        pdf_process = subprocess.run([ dvipdfmx, '-o', pdfdatei, dvidatei ], capture_output=True)
        if pdf_process.returncode == 0:
            lösche_datei(dvidatei)
            lösche_datei(auxdatei)
            lösche_datei(logdatei)
            lösche_datei(texdatei)
            print( f'    {Basisname_der_Datei}')
            Anzahl_pdf_geschrieben += 1
        else:
            print( f'{Basisname_der_Datei:<50}     pdf error {pdf_process.returncode}')
    else:
        print( f'{Basisname_der_Datei:<50}     TeX error {tex_process.returncode}')



class WatchdogHandlers(watchdog.events.FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path.endswith('.xlsx'):
            now = datetime.datetime.now().strftime('%H:%M:%S')
            time.sleep(0.2)
            fname = re.sub('^\.\/', '', event.src_path)
            print(f'    {now}    {fname} gespeichert')
            for rng in Rechnungsnummern:
                Diese_Rechnung(rng)
            if Anzahl_pdf_geschrieben == 0 and Anzahl_pdf_nicht_überschrieben > 0:
                print(f'Alle {Anzahl_pdf_nicht_überschrieben} pdf erhalten')
            print('')



########################################################
# Main
########################################################

if __name__ == "__main__":
    watchdogHandler = WatchdogHandlers()
    watchdogObserver = watchdog.observers.Observer()
    watchdogObserver.schedule(watchdogHandler, path='.', recursive=False)
    watchdogObserver.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        watchdogObserver.stop()
    watchdogObserver.join()


